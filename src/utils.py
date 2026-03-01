"""Utility functions for the Naver Smart Store - Logen integration."""

from datetime import datetime
from typing import List, Dict
from openpyxl import load_workbook


def generate_logen_filename() -> str:
    """
    Generate filename for Logen shipping Excel file with current date.
    
    The filename follows Logen's required format: "로젠발송양식_{YYYYMMDD}.xlsx"
    where YYYYMMDD is the current date.
    
    Returns:
        str: Filename in format "로젠발송양식_{YYYYMMDD}.xlsx"
        
    Example:
        >>> # If today is 2024-01-15
        >>> generate_logen_filename()
        '로젠발송양식_20240115.xlsx'
    """
    current_date = datetime.now()
    date_str = current_date.strftime("%Y%m%d")
    return f"로젠발송양식_{date_str}.xlsx"


def read_logen_excel(file_path: str) -> List[Dict[str, str]]:
    """
    Read and verify Logen shipping Excel file.
    
    Reads an Excel file in Logen format, verifies the header row matches
    the expected column order, and parses data rows starting from row 2.
    
    Args:
        file_path: Path to the Excel file to read
        
    Returns:
        List of dictionaries with keys: receiver_name, full_address, 
        receiver_tel, product_name, delivery_memo
        
    Raises:
        ValueError: If header row doesn't match expected format
        FileNotFoundError: If file doesn't exist
        
    Example:
        >>> data = read_logen_excel("로젠발송양식_20240115.xlsx")
        >>> data[0]['receiver_name']
        '홍길동'
    """
    # Expected header columns in Logen format
    expected_headers = ["받는사람", "주소", "전화번호", "상품명", "배송메모"]
    
    # Load the workbook
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Parse and verify header row (row 1)
    header_row = sheet[1]
    actual_headers = [cell.value for cell in header_row]
    
    if actual_headers != expected_headers:
        raise ValueError(
            f"Header row mismatch. Expected {expected_headers}, "
            f"but got {actual_headers}"
        )
    
    # Parse data rows starting from row 2
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(row):
            continue
            
        # Map row values to dictionary with field names
        row_data = {
            "receiver_name": row[0] if row[0] is not None else "",
            "full_address": row[1] if row[1] is not None else "",
            "receiver_tel": row[2] if row[2] is not None else "",
            "product_name": row[3] if row[3] is not None else "",
            "delivery_memo": row[4] if row[4] is not None else "",
        }
        data.append(row_data)
    
    return data
